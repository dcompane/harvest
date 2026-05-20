#!/usr/bin/env python3
"""
ctm_inventory.py

Enterprise‑scalable Control‑M Automation API inventory exporter.

FEATURES
========
- Required CLI args: --base-url, --api-key
- Optional selectors: --include deploy,config
- Metadata worksheet (timestamp, AAPI host, status, version best‑effort)
- z/OS‑aware behavior via --zos-library
- Scalable iteration:
    Servers -> Folders -> Jobs
- Excel output using openpyxl (one dataset per worksheet)

TABLES CREATED
==============
Deploy:
- Deploy Folder Job Counts
- Deploy Agent Job Counts

Config:
- Config Servers
- Per‑server Definitions (optional)
- Per‑server Agents (optional)

SAFE FOR LARGE ENVIRONMENTS
===========================
No bulk download of all jobs. All job inspection is folder‑scoped.
"""

from __future__ import annotations

import json
import sys
from dataclasses import dataclass
from datetime import datetime
from typing import Any, Dict, Optional, Iterable
from urllib.parse import urlparse

import requests
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# =============================================================================
# Authentication
# =============================================================================

@dataclass
class Auth:
    api_key: Optional[str] = None
    bearer_token: Optional[str] = None


# =============================================================================
# Control‑M Automation API Client (GET‑only)
# =============================================================================

class ControlMApi:
    def __init__(self, base_url: str, auth: Auth, timeout: int = 60):
        self.base_url = base_url.rstrip("/")
        self.timeout = timeout
        self.session = requests.Session()
        # self.session.headers["Accept"] = "application/json"
        self.session.headers["Accept"] = "*/*"
        self.session.headers["x-api-key"] = auth.api_key

    def _get(self, path: str, params: Optional[Dict[str, Any]] = None) -> Any:
        url = f"{self.base_url}{path}"
        q = {k: v for k, v in (params or {}).items() if v is not None}
        r = self.session.get(url, params=q, timeout=self.timeout, verify=False)
        r.raise_for_status()
        try:
            return r.json()
        except ValueError:
            return r.text

    # ---- Status / Metadata ----
    def get_status(self):
        return self._get("/status")

    # ---- Config ----
    def config_servers(self):
        return self._get("/config/servers")

    def config_server_definition(self, server: str):
        return self._get(f"/config/server/{server}/definition")

    def config_agents(self, server: str):
        return self._get(f"/config/server/{server}/agents")

    def config_systemsettings(self):
        return self._get("/config/systemsettings")

    # ---- Deploy ----
    def deploy_folders(self, server: str, library: Optional[str] = None):
        return self._get("/deploy/folders", params={
            "server": server,
            "library": library
        })

    def deploy_jobs(
        self,
        server: str,
        folder: str,
        library: Optional[str] = None
    ):
        return self._get("/deploy/jobs", params={
            "server": server,
            "folder": folder,
            "library": library
        })


# =============================================================================
# Excel Writer
# =============================================================================

class ExcelWorkbookWriter:
    def __init__(self):
        self.wb = Workbook()
        self.wb.remove(self.wb.active)

    def add_sheet(self, name: str, data: Any):
        ws = self.wb.create_sheet(name[:31])

        if isinstance(data, list):
            if data and isinstance(data[0], dict):
                headers = list(data[0].keys())
                ws.append(headers)
                for row in data:
                    ws.append([self._v(row.get(h)) for h in headers])
            else:
                ws.append(["Value"])
                for item in data:
                    ws.append([self._v(item)])

        elif isinstance(data, dict):
            ws.append(["Key", "Value"])
            for k, v in data.items():
                ws.append([k, self._v(v)])

        else:
            ws.append(["Value"])
            ws.append([self._v(data)])

        self._autosize(ws)

    def _v(self, val):
        if isinstance(val, (dict, list)):
            return json.dumps(val, indent=2)
        return "" if val is None else str(val)

    def _autosize(self, ws):
        for col in ws.columns:
            max_len = max(len(str(c.value)) if c.value else 0 for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 60)

    def save(self, filename: str):
        self.wb.save(filename)


# =============================================================================
# Scalable Deploy Summaries
# =============================================================================

def collect_folder_job_counts(client: ControlMApi, server: str, zos_library: Optional[str]):
    rows = []

    folders = client.deploy_folders(server=server, library=zos_library)
    if not isinstance(folders, dict):
        return rows

    for folder_path in folders.keys():
        folder = folder_path.split("/")[-1]

        try:
            deploy_data = client.deploy_jobs(server, folder, zos_library)
        except Exception as e:
            rows.append({
                "ControlMServer": server,
                "FolderName": folder,
                "SubFolderName": "",
                "JobCount": f"ERROR: {e}"
            })
            continue

        for _, folder_obj in deploy_data.items():
            jobs = folder_obj.get("Jobs", {}) or {}
            if jobs:
                rows.append({
                    "ControlMServer": server,
                    "FolderName": folder,
                    "SubFolderName": "",
                    "JobCount": len(jobs)
                })

            for sub, sub_obj in (folder_obj.get("SubFolders") or {}).items():
                sub_jobs = sub_obj.get("Jobs", {}) or {}
                rows.append({
                    "ControlMServer": server,
                    "FolderName": folder,
                    "SubFolderName": sub,
                    "JobCount": len(sub_jobs)
                })

    return rows


def collect_agent_job_counts(client: ControlMApi, server: str, zos_library: Optional[str]):
    agent_counts: Dict[str, int] = {}

    folders = client.deploy_folders(server=server, library=zos_library)
    if not isinstance(folders, dict):
        return []

    for folder_path in folders.keys():
        folder = folder_path.split("/")[-1]

        try:
            deploy_data = client.deploy_jobs(server, folder, zos_library)
        except Exception:
            continue

        for _, folder_obj in deploy_data.items():
            jobs = folder_obj.get("Jobs", {}) or {}
            for _, job in jobs.items():
                agent = (
                    job.get("Agent")
                    or job.get("Host")
                    or job.get("Node")
                    or "UNASSIGNED"
                )
                agent_counts[agent] = agent_counts.get(agent, 0) + 1

    return [
        {"ControlMServer": server, "Agent": agent, "JobCount": count}
        for agent, count in sorted(agent_counts.items())
    ]


# =============================================================================
# Utility
# =============================================================================

def parse_include(val: str):
    allowed = {"deploy", "config", "auth", "provision"}
    parts = {p.strip().lower() for p in val.split(",") if p.strip()}
    unknown = parts - allowed
    if unknown:
        raise ValueError(f"Invalid --include value(s): {','.join(unknown)}")
    return parts or allowed


# =============================================================================
# Main
# =============================================================================

def main(argv: Optional[Iterable[str]] = None):
    import argparse

    parser = argparse.ArgumentParser(
        description="Scalable Control‑M Automation API inventory exporter",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter
    )
    parser.add_argument("--base-url", required=True)
    parser.add_argument("--api-key", required=True)
    parser.add_argument("--include", default="deploy,config")
    parser.add_argument("--zos-library")
    parser.add_argument("--timeout", type=int, default=60)
    parser.add_argument("--output", default=f"controlm_inventory_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
    parser.add_argument("--server", action="append")
    parser.add_argument("--debug", default=False)

    args = parser.parse_args(list(argv) if argv else None)

    includes = parse_include(args.include)

    client = ControlMApi(args.base_url, Auth(api_key=args.api_key), args.timeout)
    excel = ExcelWorkbookWriter()

    # ---- Metadata ----
    parsed = urlparse(args.base_url)
    metadata = {
        "Timestamp": datetime.now().isoformat(),
        "AAPI Base URL": args.base_url,
        "AAPI Host": parsed.hostname,
        "AAPI Status": client.get_status()
    }
    excel.add_sheet("Metadata", metadata)

    servers = client.config_servers()
    server_names = [
        s["name"] if isinstance(s, dict) else s
        for s in servers
    ]

    # ---- Config ----
    if "config" in includes:
        excel.add_sheet("Config Servers", servers)

        for srv in args.server or []:
            excel.add_sheet(f"{srv} Definition", client.config_server_definition(srv))
            excel.add_sheet(f"{srv} Agents", client.config_agents(srv))

    # ---- Deploy ----
    if "deploy" in includes:
        folder_rows = []
        agent_rows = []

        for srv in server_names:
            folder_rows.extend(
                collect_folder_job_counts(client, srv, args.zos_library)
            )
            agent_rows.extend(
                collect_agent_job_counts(client, srv, args.zos_library)
            )

        excel.add_sheet("Deploy Folder Job Counts", folder_rows)
        excel.add_sheet("Deploy Agent Job Counts", agent_rows)

    excel.save(args.output)
    print(f"\n✅ Inventory workbook created: {args.output}")


if __name__ == "__main__":
    sys.exit(main())