# =============================================================================
# Excel Writer
# =============================================================================
import os
import platform
from typing import Any, Optional
import json

from click import style
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

class ExcelWorkbookWriter:
    """Create Excel workbooks with optional Excel Tables using openpyxl."""

    def __init__(self):
        self.wb = Workbook()
        self.wb.remove(self.wb.active)
        self._table_counter = 0

    # ------------------------------------------------------------------ #
    # PUBLIC API
    # ------------------------------------------------------------------ #

    def add_sheet(self, sheet_name: str):
        """Create a new sheet and add data (as a table by default)."""

        return self.wb.create_sheet(sheet_name[:31])

    def add_table(
        self,
        sheet_name: str,
        data: Any,
        start_row: Optional[int] = None,
        start_col: Optional[int] = 1,
        as_table: bool = True,
        table_name: Optional[str] = None,
        direction: Optional[str] = "horizontal",
        gap_rows: int =2,
        table_title: str = None,
        description: str = None,
        columns: Optional[list] = None,
        index: Optional[int] = None,
        tab_color: Optional[str] = None
    ):
        """
        Add a dataset to an existing sheet as a separate table.

        If start_row is not provided, data is appended below existing content.
        """
        style = "TableStyleMedium10"
        

        if sheet_name not in self.wb.sheetnames:
            # index must be always positive. If not, create as last sheet
            if index is not None and index >= 0:
                # Otherwise, insert at index position (0-based) in the workbook's sheet list
                ws = self.wb.create_sheet(sheet_name[:31], index=index)
            else:
                ws = self.wb.create_sheet(sheet_name[:31])

            ws.sheet_properties.tabColor = tab_color if tab_color else None
        else:
            ws = self.wb[sheet_name]

        if start_row is None:
            if ws.max_row > 1:
                for _ in range(gap_rows):
                    ws.append([])  # Add (gap_rows) empty rows for spacing
                start_row = ws.max_row + gap_rows  # Start after the gap
            else:
                start_row = ws.max_row

        if table_title:
            ws.append([f"Table title: {table_title}"])
        if description:
            ws.append([f"Description: {description}"])
        if table_title is not None or description is not None:
            ws.append([""])  # Add an empty row for spacing after title/description
            start_row = ws.max_row + 1  # Update start_row after adding title/description

        headers, rows = self._normalize_data(data,direction=direction, columns=columns)

        if columns:
            for i in range(max(len(headers), len(columns))):
                if i < min(len(headers), len(columns)):
                    headers[i] = columns[i]
                elif i >= len(headers) and i <= len(columns):
                    headers.append(columns[i])

        # Write headers
        ws.append(headers)
        header_row = start_row 

        if any(isinstance(item, list) for item in rows):
            # Single row of data - treat headers as keys and values as a single row
            nested_rows = rows
        else:
            nested_rows = [rows] # Ensure it's a list of rows, even if there's only one

            # Write rows
        for row in nested_rows:
            ws.append(self._convert_types(row))

        end_row = header_row + len(nested_rows)
        end_col = len(headers)

        if as_table:
            self._add_excel_table(
                ws=ws,
                start_row=header_row,
                start_col=start_col,
                end_row=end_row,
                end_col=end_col,
                table_name=table_name,
                style=style
            )

        self._autosize(ws)

    def save(self, filename: str):
        """Save the workbook to a file."""
        self.wb.save(filename)

    # ------------------------------------------------------------------ #
    # INTERNAL HELPERS
    # ------------------------------------------------------------------ #

    def _remove_newlines(self, obj):
        """Recursively remove newlines from strings in a nested structure."""
        if isinstance(obj, str):
            return obj.replace("\n", "")
        elif isinstance(obj, list):
            return [self._remove_newlines(item) for item in obj]
        elif isinstance(obj, tuple):
            return tuple(self._remove_newlines(item) for item in obj)
        elif isinstance(obj, dict):
            return {k: self._remove_newlines(v) for k, v in obj.items()}
        return obj  # Leave other types unchanged

    def _normalize_data(self, data: Any, direction: str = "horizontal", columns: Optional[list] = None):
        """Normalize input data into headers + rows."""

        # Remove New Lines from all string values in the data structure
        data = self._remove_newlines(data)

        if isinstance(data, list) and data and isinstance(data[0], dict):
            if columns:
                headers = columns
            else:
                headers = list(data[0].keys())
            for i in range(1, len(data)):
                for h in  list(data[i].keys()):
                    if h not in headers:
                        headers.append(h)

            rows = [[self._v(row.get(h)) for h in headers] for row in data]

        elif isinstance(data, list):
            headers = ["Value"]
            rows = [[self._v(item)] for item in data]

        elif isinstance(data, dict) and direction == "horizontal":
            headers = list(data.keys())
            rows = [v for k, v in data.items()]

        elif isinstance(data, dict) and direction == "vertical":
                headers = ["Key", "Value"]
                rows = [[k, self._v(v)] for k, v in data.items()]

        else:
            headers = ["Value"]
            rows = [[self._v(data)]]

        return headers, rows

    def _add_excel_table(
        self,
        ws,
        start_row,
        start_col,
        end_row,
        end_col,
        table_name,
        style
    ):
        # Generate default table name: Table1, Table2, ...
        self._table_counter += 1
        table_name = table_name or f"Table{self._table_counter}"

        ref = (
            f"{get_column_letter(start_col)}{start_row}:"
            f"{get_column_letter(end_col)}{end_row}"
        )

        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name=style,
            showRowStripes=True,
            showColumnStripes=False,
        )

        ws.add_table(table)

    def _v(self, val):
        if isinstance(val, (dict, list)):
            return json.dumps(val, indent=2)
        return "" if val is None else str(val)

    def _autosize(self, ws):
        for col in ws.columns:
            max_len = max(len(str(c.value)) if c.value else 0 for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(
                max_len + 2, 60
            )

    # Convert numeric-looking strings to actual numbers
    def _convert_types(self,row):
        converted = []
        for value in row:
            # Try converting to int or float
            if isinstance(value, str):
                try:
                    if "." in value:
                        converted.append(float(value))
                    else:
                        converted.append(int(value))
                except ValueError:
                    converted.append(value)  # Keep as string if not numeric
            else:
                converted.append(value)
        return converted

# =============================================================================
# Main
# =============================================================================
if __name__ == "__main__":
    assert "You should not see this message" == "This file is not meant to be run directly."